#!/bin/python  
# Generates a collection of dummy social media data  
from random import choice, randint, random  
from time import strftime  
from datetime import timedelta, datetime  
from openpyxl import Workbook  
import xlrd  
#Reads lines "NIKE 23 14 45" from 7days.xlsx which is the count of pos, neg and neu posts to be generated for NIKE in the given period  
def get_products_and_senti_num():  
    book = xlrd.open_workbook('senti_count_total.xlsx')  
    sh = book.sheet_by_index(0)  
    products = []  
    senti_num = []  
    for rownum in range(sh.nrows):  
        products.append(sh.row_values(rownum)[0])  
        senti_num.append(sh.row_values(rownum)[1:4])  
    return products, senti_num  
#Returns prefix + ndigits  
def randomN(prefix, ndigits):  
    range_start = 10**(ndigits-1)  
    range_end = (10**ndigits)-1  
    return prefix + str(randint(range_start, range_end))  
def random_date(start, end):  
    return start + timedelta(  
        seconds=randint(0, int((end - start).total_seconds())))  
def gen_posts(s_date, e_date):  
    social_book = Workbook(optimized_write = True)  
    social_sheet = social_book.create_sheet()  
    voice_book = Workbook(optimized_write = True)  
    voice_sheet = voice_book.create_sheet()  
    start_datetime = datetime(s_date[2], s_date[1], s_date[0], 0, 0, 0)  
    end_datetime = datetime(e_date[2], e_date[1], e_date[0] + 1, 0, 0, 0)  
    client_list = ['001']  
    user_list = ['John', 'William', 'James', 'Jacob', 'Ryan', 'Joshua', 'Michael', 'Jayden', 'Ethan', 'Christopher', 'Samuel', 'Daniel', 'Kevin', 'Elijah']  
    channel_list = ['TW', 'FB']  
    countries = ['India', 'Germany', 'France', 'The United States']  
    locations = {"India" : ["Bangalore", "Chennai", "Delhi", "Mumbai"],  
                "Germany": ["Berlin", "Munich", "Stuttgart", "Frankfurt"],  
                "France": ["Paris", "Marseille", "Lyon"],  
                "The United States": ["Florida", "Washington DC", "Texas", "Dallas"]}  
    country_codes = {"India": "IN",  
                    "Germany" : "DE",  
                    "France" : "FR",  
                    "The United States": "US"}  
#The adj_set has the adjectives that will be used in the posts.  
    adj_set = {"good" : ['good', 'nice'],  
          "very_good" : ['refreshing', 'magical'],  
          "neutral" : ['ok'],  
          "bad" : ['not good', 'substandard', 'unpleasant', 'poor'],  
          "very_bad" : ['awful', 'horrible', 'terrible']}  
    adj_kind_from_senti = { 2 : "very_good",  
                1 : "good",  
                0 : "neutral",  
                -1 : "bad",  
                -2 : "very_bad"}  
    post_templates = {"very_good" : ["Hey guys, try {0}, it is {1}! Dont miss!",  
                      "People, I got the new {0} - {1}!! Brilliant! Give a try!",  
                      "I'm loving {0}!!",  
                      "Using {0} feels great!!",  
                      "{0} is {1}. My body feels so refreshing",  
                      "{0} - The product quality is impressive!! Verdict - {1}",  
                      "{0} is {1}. Highly recommended",  
                      "{0} gives instant refreshing moisturizing effect!"],  
            "good"      : ["Today I tried {0}. It is {1}.",  
                            "The new {0}. Product quality is top, is {1} and worth a try",  
                            "Did you checkout {0}?, {1} thing.",  
                            "I like {0}. It smells nice and so soft",  
                            "Didnt know {0} is {1} stuff. Superb!. Do try it."],  
            "neutral"  : ["Checked out {0}. It is {1}",  
                            "The new {0} is {1}. Dont expect much.",  
                            "Heard the new {0} is {1}. Any first hand info on the it?",  
                            "Anyone know how is {0}, reviews say it is {1}. Quality is what matters"],  
            "very_bad"  : ["OMG!! Tried {0}. Its not for you. It is {1}",  
                            "Never go for {0}, the quality is very less, {1} thing.",  
                            "Oh, such a {1} thing {0} is!",  
                            "{0} is sold out in my area - Sad!",  
                            "Couldnt find {0} in my local store. Bad that I cant get that.",  
                            "Local stored have sold out {0}, please send in more!!",  
                            "We need more stock of {0} in here. Out of stock everywhere I check",  
                            "{0} is out of stock - So sad!",  
                            "Dont ever think of getting a {0}, very bad product. It is {1}",  
                            "Why do we have {1} products like {0}? :("],  
            "bad"      : ["Tried the new {0}. It is not recommended - {1}",  
                            "Shouldnt have gone for the {1} {0}. Pathetic product quality.",  
                            "First hand experience: {0} is {1}!",  
                            "10 stores and no {0}. I want it desperately",  
                            "Tried finding {0}. Can't find it in any stores in my area.",  
                            "My {0} is {1}. The quality is way too less. Is it just me?!",  
                            "The new {0} is {1}. It is disappointing. Fail!!"]}  
    products, senti_num = get_products_and_senti_num()  
    for j in range(len(products)):  
        product = products.pop()  
        senti = senti_num.pop()  
        pos = int(senti[0])  
        neg = int(senti[1])  
        neu = int(senti[2])  
        print product, "-", pos, neg, neu, " posts created."  
        for k in range(pos + neg + neu):  
            if pos:  
                sentiment = randint(1,2)  
                pos -= 1  
            elif neg:  
                sentiment = randint(-2,-1)  
                neg -= 1  
            else:  
                sentiment = 0  
                neu -= 1  
            sentiment_valuation = sentiment + 3 if sentiment else sentiment  
            adj_kind = adj_kind_from_senti[sentiment]  
            adj = choice(adj_set[adj_kind])  
            client = choice(client_list)  
            guid = randomN('POB', 29)  
            user = choice(user_list)  
            channel = choice(channel_list)  
            post_template = choice(post_templates[adj_kind])  
            posted_on = random_date(start_datetime, end_datetime)  
            post = post_template.format(product, adj)  
            num_of_votes = str(randint(0, 150))  
            if channel == 'TW':  
                post_link = 'http://twitter.com/' + user + randomN('', 5)  
            if channel == 'FB':           
                post_link = 'http://facebook.com/' + user + randomN('', 5)  
            post_type = choice(['Status', 'Link', 'Photo', 'Video'])  
            country = choice(countries)  
            location = choice(locations[country])  
            country_code = country_codes[country]  
            latitude = str(randomN("", 2) + '.' + str(randint(2, 20)))  
            longitude = str(randomN("", 2) + '.' + str(randint(2, 20)))  
            social_sheet.append([client, guid, channel[:2].upper() + str(randomN('',6)), 'English', channel, user, posted_on.strftime("%a, %d %b %Y %H:%M:%S +0000"), post_type, post_link, num_of_votes, location, country, latitude, longitude, '3', 'Demo post', user, 'Demo User Retrieval', product, posted_on.strftime("%Y%m%d%H%M%S"), post, posted_on.strftime("%Y%m%d%H%M%S"), 'Demo Post Parent', "DemoJ", country_code, 'DS'])  
            voice_sheet.append([client, guid, 'TextAnalysis', 'Sentiment', 'DEMO', sentiment, sentiment_valuation, 'J', posted_on.strftime("%Y%m%d"), posted_on.strftime("%Y%m%d%H%M%S")])  
            voice_sheet.append([client, guid, 'TextAnalysis', 'PRODUCT', product, sentiment, sentiment_valuation, 'J', posted_on.strftime("%Y%m%d"), posted_on.strftime("%Y%m%d%H%M%S")])  
    social_book.save('SOCIALDATA.xlsx')  
    voice_book.save('SMI_VOICE_CUST.xlsx')  
    print 'Demo data saved in SOCIALDATA.xlsx, SMI_VOICE_CUST.xlsx'  
#modify this line => gen_posts(start_date, end_date)  
gen_posts([22, 05, 2014], [05, 06, 2014]) 